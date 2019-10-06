import os
import openpyxl
import logging
logging.basicConfig(level=logging.DEBUG, format="[%(asctime)s %(funcName)s] %(message)s")
log = logging.getLogger(__name__)

log.debug("start, excel");
path = os.path.join(os.path.dirname(__file__),"../assets","app_memo.xlsx")

class Memo:
    wb=''
    def __init__(self):
        self.wb = openpyxl.load_workbook(path, data_only=True)

    def getAllData(self):
        tmp = list()
        ws = self.wb.get_sheet_by_name("main")
        for row in ws.rows:
            index = row[0].value
            if type(index) == int:
                index, category, word, meaning = row[0].value, row[1].value, row[2].value, row[3].value
                log.debug("1.{},{}".format(word, meaning))

                examples, keywords = self.getData( index )

                tmp.append( dict(index=index, category=category, word=word, meaning=meaning, examples=examples, keywords=keywords) )

        # log.debug("Excel data: %s" % str(tmp) )
        return tmp

    def clear(self):
        self.wb=''
    
    def remove(self, list_data):
        if not list_data:
            return
        self.key_index = int( list_data['index'] )
        log.debug("index:%s"%self.key_index)
        for key in ['main','examples','keywords']:
            self.removeData(key, [])

    def removeElement(self, list_data):
        if not list_data:
            return
            
        for item in list_data:
            self.key_index = int( item['main']['index'] )
            log.debug("key index %s"%self.key_index)
            for (sname, data) in item.items():
                log.debug("sheetname:%s"%sname)
                self.removeData(sname, data)

    def updateElement(self, list_data):
        if not list_data:
            return
            
        for item in list_data:
            self.key_index = int( item['main']['index'] )
            log.debug("key index %s"%self.key_index)
            for (sname, data) in item.items():
                log.debug("sheetname:%s"%sname)
                self.updateData(sname, data)

    def addElement(self, list_data):
        if not list_data:
            return

        for item in list_data:
            self.key_index = int( item['index'] )
            log.debug(self.key_index)
            for p in (p for p in item if p!='index'):
                self.addData(p, item[p])

        self.wb.save(path)

    def add(self, list_data):
        if not list_data:
            return
        data = list_data[0]
        # index 지정
        ws = self.wb.get_sheet_by_name('main')
        self.key_index = ws.cell(row=ws.max_row, column=1).value +1
        self.new_examples = len(data.get('examples'))
        self.new_keywords = len(data.get('keywords'))

        log.debug("DATA:{}".format(data))

        for k,v in data.items():
            sheetname = k
            self.addData(sheetname, v)
        
        self.wb.save(path)

    @staticmethod
    def getLastRow(ws):
        for ix, row in enumerate(ws.rows,1):
            if not row[1].value:
                return ix -1
        return ix

    def removeData(self, sn, data):
        ws = self.wb.get_sheet_by_name(sn)
        log.debug("data %s"%data)

        if(sn =='main'):
            # 데이터가 있으면 반환, 없는 경우한해 메인포함 전체삭제
            if data:
                return;

            # 시트명, 찾을 컬럼위치(0부터시작)
            row_no = self.find_row_no(ws,0)
            log.debug("row result %s"%row_no)
            if row_no == -1:
                return -1

            ws.delete_rows(row_no, 1)

        else:
            if data:
                for index in sorted(data['index'], reverse=True):
                    row_no = self.find_row_no(ws, 1, nth=int(index) )
                    log.debug("index_no:%d"%index)
                    log.debug("row result %s"%row_no)
                    if row_no == -1:
                        return -1

                    ws.delete_rows(row_no, 1)
            else:
                row_no = self.find_row_no(ws, 1, nth=0)
                while row_no!=-1:
                    ws.delete_rows(row_no, 1)
                    row_no = self.find_row_no(ws, 1, nth=0)

        self.wb.save(path)

    def updateData(self, sn, data):
        ws = self.wb.get_sheet_by_name(sn)
        if(sn =='main'):
            # 시트명, 찾을 컬럼위치(0부터시작)
            row_no = self.find_row_no(ws,0)
            log.debug("row result %s"%row_no)
            if row_no == -1:
                return -1

            schema = {'category':2, 'word':3, 'meaning':4}
            for p in data:
                if p in schema:
                    ws.cell(row=row_no, column=schema[p]).value = data[p]
                    log.debug(data[p])
        else:
            for index, data in zip(data['index'], data['data']):
                row_no = self.find_row_no(ws, 1, nth=int(index) )
                log.debug("index_no:%d"%index)
                log.debug("row result %s"%row_no)
                if row_no == -1:
                    return -1

                ws.cell(row=row_no, column=3).value = data
                log.debug(data)

        self.wb.save(path)

    # worksheetname, 찾을 column번호, 몇번째 데이타인지(기본 첫번쨰)
    def find_row_no(self, ws, column_no, nth=0):
        count = 0
        for row_no, row in enumerate(ws.rows, start=1):
            log.debug("process {}=={},{}=={}".format(count, nth, row[column_no].value, self.key_index))
            if row[column_no].value == self.key_index:
                if count == nth:
                    log.debug("I got it")
                    return row_no
                count=count + 1
        return -1

    def addData(self, sn, data):
        ws = self.wb.get_sheet_by_name(sn)
        # new_row_no = ws.max_row+1
        new_row_no = self.getLastRow(ws) + 1

        if sn=='main':
            ws.cell(row=new_row_no, column=1).value = self.key_index
            # category
            ws.cell(row=new_row_no, column=2).value = data[0]
            # word
            ws.cell(row=new_row_no, column=3).value = data[1]
            # meaning
            ws.cell(row=new_row_no, column=4).value = data[2]
            # examples count
            ws.cell(row=new_row_no, column=5).value = self.new_examples
            # keyword count
            ws.cell(row=new_row_no, column=6).value = self.new_keywords
        else:
            # 마지막 인덱스값을 취득
            last_index = ws.cell(row=new_row_no -1, column=1).value
            for ix, v in enumerate(data,1):
                ws.cell(row=new_row_no + ix -1, column=1).value = last_index + ix
                ws.cell(row=new_row_no + ix -1, column=2).value = self.key_index
                ws.cell(row=new_row_no + ix -1, column=3).value = v
                
        log.debug("{} {}".format(sn, data))

    def getCategory(self):
        ws = self.wb.get_sheet_by_name("main")
        cat_list = set()
        for row_no in range(2,ws.max_row+1,1):
            cat_list.add(ws.cell(row=row_no, column=2).value)

        return list(cat_list)

    def getData(self, key_index):
        ws = self.wb.get_sheet_by_name("examples")
        examples = list()
        for row in ws.rows:
            if row[1].value == key_index:
                examples.append(row[2].value)
        
        log.debug("2.{}".format(str(examples)))
        
        ws = self.wb.get_sheet_by_name("keywords")
        keywords = list()
        for row in ws.rows:
            if row[1].value == key_index:
                keywords.append(row[2].value)

        log.debug("3.{}".format(str(keywords)))

        return examples, keywords

# 1부터 시작되는 행수반환
def getRow(worksheet, search_target):
    for index, row in enumerate(worksheet.rows):
        if row[0].value == search_target:
            return index+1
    return None

def exam1():
    ws = wb.get_active_sheet()
    for row in ws.rows:
        for cell in row:
            log.debug(cell.value)

    # for column in ws.columns:
    #     for cell in column:
    #         log.debug(cell.value)

def exam2():    
    ws = wb.get_active_sheet() 
    ws['A3'] = 100
    ws['B3'].select
    # log.debug("activecell %s" % activecell.value)
    path = os.path.join(os.path.dirname(__file__),"../assets","test.xlsx")
    wb.save(filename=path)
    # exam1()

def exam3():    
    ws = wb.get_active_sheet() 
    ws['A3'] = 100
    log.debug("cell 1,1 %s" % ws.cell(1,1).value )
    row_no =1
    while ws.cell(row_no,1).value:
        log.debug("row %d: %s" %(row_no, ws.cell(row_no,1).value) )
        row_no = row_no +1
    
    # log.debug("activecell %s" % activecell.value)

if __name__=='__main__':
    getAllData()