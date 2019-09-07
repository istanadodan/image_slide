import os
import openpyxl
import logging
logging.basicConfig(level=logging.DEBUG, format="[%(asctime)s %(funcName)s] %(message)s")
log = logging.getLogger(__name__)

log.debug("start, excel");
path = os.path.join(os.path.dirname(__file__),"../assets","app_memo.xlsx")
wb = openpyxl.load_workbook(path, data_only=True)

def getData(key_index):
    ws = wb.get_sheet_by_name("main")
    rows_no = getRow(ws, key_index)
    word, meaning = ws.cell(rows_no,3), ws.cell(rows_no,4)
    log.debug("1.{},{}".format(word.value, meaning.value))

    ws = wb.get_sheet_by_name("examples")
    examples = list()
    for row in ws.rows:
        if row[1].value == key_index:
            examples.append(row[2].value)
    
    log.debug("2.{}".format(str(examples)))
    
    ws = wb.get_sheet_by_name("keywords")
    keywords = list()
    for row in ws.rows:
        if row[1].value == key_index:
            keywords.append(row[2].value)

    log.debug("3.{}".format(str(keywords)))

    return word.value, meaning.value, examples, keywords

# 1부터 시작되는 행수반환
def getRow(worksheet, search_target):
    for index, row in enumerate(worksheet.rows):
        if row[0].value == search_target:
            return index+1
    return None

def exam1():
    ws = wb.get_active_sheet()
    for row in ws.rows:
        for cell in row:
            log.debug(cell.value)

    # for column in ws.columns:
    #     for cell in column:
    #         log.debug(cell.value)

def exam2():    
    ws = wb.get_active_sheet() 
    ws['A3'] = 100
    ws['B3'].select
    # log.debug("activecell %s" % activecell.value)
    path = os.path.join(os.path.dirname(__file__),"../assets","test.xlsx")
    wb.save(filename=path)
    # exam1()

def exam3():    
    ws = wb.get_active_sheet() 
    ws['A3'] = 100
    log.debug("cell 1,1 %s" % ws.cell(1,1).value )
    row_no =1
    while ws.cell(row_no,1).value:
        log.debug("row %d: %s" %(row_no, ws.cell(row_no,1).value) )
        row_no = row_no +1
    
    # log.debug("activecell %s" % activecell.value)

if __name__=='__main__':
    getData(1)
    getData(2)